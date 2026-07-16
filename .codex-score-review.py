import json, re, sys
from pathlib import Path
import requests
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

BOOK = Path(r"C:\GitHub\scores\第一階段書審評分表-蔡芸琤.xlsx")
OUT = Path(r"C:\GitHub\scores\.review_downloads")
OUT.mkdir(exist_ok=True)

def formula_url(value):
    if not isinstance(value, str):
        return None
    m = re.search(r'HYPERLINK\("([^"]+)', value, re.I)
    return m.group(1) if m else None

def drive_id(url):
    for pat in (r'[?&]id=([^&]+)', r'/d/([^/]+)'):
        m = re.search(pat, url)
        if m:
            return m.group(1)
    return None

s = requests.Session()
s.headers['User-Agent'] = 'Mozilla/5.0'
wb = load_workbook(BOOK, data_only=False)
results = []
for sheet_name in ('大專組', '高中職組'):
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row, 3).value:
            continue
        case = str(ws.cell(row, 3).value)
        pdf_url = formula_url(ws.cell(row, 4).value)
        video_url = formula_url(ws.cell(row, 5).value)
        rec = {'sheet': sheet_name, 'row': row, 'team': ws.cell(row,1).value,
               'title': ws.cell(row,2).value, 'case': case,
               'pdf_url': pdf_url, 'video_url': video_url}
        try:
            fid = drive_id(pdf_url or '')
            u = f'https://drive.google.com/uc?export=download&id={fid}&confirm=t'
            r = s.get(u, timeout=90)
            rec['pdf_http'] = r.status_code
            rec['pdf_type'] = r.headers.get('content-type','')
            rec['pdf_bytes'] = len(r.content)
            if r.ok and r.content.startswith(b'%PDF'):
                p = OUT / f'{case}.pdf'
                p.write_bytes(r.content)
                rec['pdf_path'] = str(p)
                rec['pdf_ok'] = True
            else:
                rec['pdf_ok'] = False
                (OUT / f'{case}.download').write_bytes(r.content)
        except Exception as e:
            rec['pdf_ok'] = False
            rec['pdf_error'] = repr(e)
        try:
            if 'youtu' in (video_url or ''):
                ou = 'https://www.youtube.com/oembed?url=' + requests.utils.quote(video_url, safe='') + '&format=json'
                vr = s.get(ou, timeout=30)
            else:
                vr = s.get(video_url, timeout=45, allow_redirects=True)
            rec['video_http'] = vr.status_code
            rec['video_final'] = vr.url
            rec['video_ok'] = vr.ok
        except Exception as e:
            rec['video_ok'] = False
            rec['video_error'] = repr(e)
        results.append(rec)
        print(json.dumps(rec, ensure_ascii=False), flush=True)
(OUT / 'links.json').write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding='utf-8')
