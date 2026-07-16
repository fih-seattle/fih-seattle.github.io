import csv, html, io, json, math, os, re, statistics, urllib.request
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw, ImageFont

ROOT=Path(r'C:\GitHub\high-school\114')
WORK=Path(r'C:\GitHub\fih-seattle.github.io')
ASSETS=WORK/'.codex_report_assets'
ASSETS.mkdir(exist_ok=True)

font_path=Path(r'C:\Windows\Fonts\msjh.ttc')
FONT=ImageFont.truetype(str(font_path),22) if font_path.exists() else ImageFont.load_default()
FONT_SM=ImageFont.truetype(str(font_path),16) if font_path.exists() else ImageFont.load_default()
FONT_LG=ImageFont.truetype(str(font_path),30) if font_path.exists() else ImageFont.load_default()

def rows(ws): return [r for r in ws.iter_rows(values_only=True) if any(v not in (None,'') for v in r)]
def pct(a,b): return 100*a/b if b else 0
def mean(x): return statistics.mean(x) if x else 0
def med(x): return statistics.median(x) if x else 0
def fmt(x,n=1): return f'{x:.{n}f}'
def esc(x): return html.escape(str(x))

# Locate core files without relying on locale-sensitive literals.
xlsx=list(ROOT.glob('*.xlsx'))
grade_path=next(p for p in xlsx if p.name.startswith('114-1') and p.stat().st_size>100000)
roster_path=next(p for p in xlsx if p.name.startswith('114-2 ') and p.stat().st_size>100000)
feedback_path=next(p for p in xlsx if p.name.startswith('@@-0119'))

# 114-1 assessment evidence.
wb=openpyxl.load_workbook(grade_path,data_only=True,read_only=True)
assessment=[]
for ws in wb.worksheets:
    rr=rows(ws); hdr=rr[0]
    is_alt='_' in ws.title
    score_idx=2 if len(hdr)>2 and str(hdr[2]).strip() in ('成績','Score') else 1
    vals=[]
    for r in rr[1:]:
        if len(r)>score_idx:
            try: vals.append(float(r[score_idx]))
            except: pass
    base=ws.title.split('_')[0]
    rec=next((x for x in assessment if x['name']==base),None)
    if not rec:
        rec={'name':base,'scores':[],'regular':0,'alt':0,'statuses':Counter()}; assessment.append(rec)
    rec['scores']+=vals
    rec['alt' if is_alt else 'regular']+=len(vals)
    status_idx=3 if score_idx==2 else 2
    for r in rr[1:]:
        if len(r)>status_idx and r[status_idx] not in (None,''): rec['statuses'][str(r[status_idx]).strip()]+=1
for a in assessment:
    s=a.pop('scores'); a.update(n=len(s),avg=mean(s),median=med(s),ge90=sum(x>=90 for x in s),ge60=sum(x>=60 for x in s),perfect=sum(x==100 for x in s))
total_1141=sum(a['n'] for a in assessment)
alt_1141=sum(a['alt'] for a in assessment)

# 114-1 feedback.
fb=openpyxl.load_workbook(feedback_path,data_only=True,read_only=True)
qrows=rows(fb.worksheets[0]); qhdr=qrows[0]; n_feedback=len(qrows)-1
likert=[]
for j in range(5,len(qhdr)):
    vals=[]
    for r in qrows[1:]:
        try: vals.append(float(r[j]))
        except: pass
    likert.append({'q':str(qhdr[j]),'mean':mean(vals),'positive':pct(sum(v>=4 for v in vals),len(vals)),'n':len(vals)})
qual=rows(fb.worksheets[1]); qlh=qual[0]
interaction=Counter(str(r[7]).strip() for r in qual[1:] if len(r)>7 and r[7] not in (None,''))
post_freq=Counter(str(r[6]).strip() for r in qual[1:] if len(r)>6 and r[6] not in (None,''))

# 114-2 roster and weekly worksheets.
wb2=openpyxl.load_workbook(roster_path,data_only=True,read_only=True)
roster=rows(wb2.worksheets[1]); rhead=roster[0]; students=len(roster)-1
schools=len(set(str(r[1]).strip() for r in roster[1:] if len(r)>1 and r[1] not in (None,'')))
weekly=[]
completion_by_student=[]
for r in roster[1:]: completion_by_student.append(sum(1 for j in range(4,14) if len(r)>j and str(r[j]).strip().upper()=='V'))
for j in range(4,14):
    n=sum(1 for r in roster[1:] if len(r)>j and str(r[j]).strip().upper()=='V')
    label=rhead[j].strftime('%m/%d') if hasattr(rhead[j],'strftime') else str(rhead[j])
    weekly.append({'week':label,'n':n,'rate':pct(n,students)})
selfweeks=[]; all_ratings=Counter()
for ws in wb2.worksheets[3:13]:
    rr=rows(ws); hdr=rr[0]
    idx=next((i for i,v in enumerate(hdr) if v and '1)' in str(v)),None)
    c=Counter(str(r[idx]).strip() for r in rr[1:] if idx is not None and len(r)>idx and r[idx] not in (None,''))
    all_ratings.update(c)
    selfweeks.append({'week':ws.title.split('(')[0],'n':len(rr)-1,'ratings':dict(c),
                      'understand_rate':pct(c.get('大致懂',0)+c.get('很懂',0),sum(c.values()))})

# Final project comment corpus.
txt=(ROOT/'New Text Document.txt').read_text(encoding='utf-8',errors='replace')
blocks=re.split(r'(?=^[A-Za-z0-9][\w-]* commented on )',txt,flags=re.M)
blocks=[b for b in blocks if 'commented on ' in b]
project={'comment_blocks':len(blocks),'with_colab':0,'with_video':0,'with_both':0,'with_docs':0}
for b in blocks:
    c='colab.research.google.com' in b; v=('youtu.be/' in b or 'youtube.com/' in b)
    project['with_colab']+=c; project['with_video']+=v; project['with_both']+=(c and v); project['with_docs']+='docs.google.com' in b

# Public peer-review sheet export.
peer_url='https://docs.google.com/spreadsheets/d/1UJTLNFdzU5qxodwN4okBHZblVoWy6vIc_JVtK_MwCOE/export?format=csv&gid=992136598'
peer_bytes=urllib.request.urlopen(urllib.request.Request(peer_url,headers={'User-Agent':'Mozilla/5.0'}),timeout=30).read()
peer_rows=list(csv.reader(io.StringIO(peer_bytes.decode('utf-8-sig'))))
ph=peer_rows[0]; pdata=[r for r in peer_rows[1:] if any(x.strip() for x in r)]
score_cols=[i for i,h in enumerate(ph) if '互評分數' in h]
peer_scores=[]
for r in pdata:
    for i in score_cols:
        if i<len(r):
            try: peer_scores.append(float(r[i]))
            except: pass
reflection_cols=[i for i,h in enumerate(ph) if any(k in h for k in ['我從別人的作品學到','如果我重做一次','最大困難','給下一屆'])]
reflection_complete=sum(1 for r in pdata if all(i<len(r) and r[i].strip() for i in reflection_cols))

# Link audit.
audit=json.load(open(WORK/'.codex_high_school_114_link_audit.json',encoding='utf-8'))['summary']

# Charts drawn with Pillow to avoid external plotting dependencies.
blue='#1f4e79'; teal='#2a8c82'; orange='#e09132'; gray='#64748b'
def bar_chart(labels, values, title, ylabel, path, colors=None, ymax=None):
    W,H=1600,720; L,R,T,B=140,50,90,165
    im=Image.new('RGB',(W,H),'white'); d=ImageDraw.Draw(im)
    ymax=ymax or max(values)*1.15 or 1
    for k in range(6):
        y=T+(H-T-B)*k/5; val=ymax*(5-k)/5
        d.line((L,y,W-R,y),fill='#d7e0e7',width=1); d.text((15,y-12),f'{val:.0f}',font=FONT_SM,fill=gray)
    n=len(values); slot=(W-L-R)/n; bw=slot*.66
    for i,(lab,val) in enumerate(zip(labels,values)):
        x=L+i*slot+(slot-bw)/2; y=H-B-(H-T-B)*val/ymax
        d.rectangle((x,y,x+bw,H-B),fill=(colors[i] if colors else blue))
        d.text((x+bw/2,y-30),f'{val:.1f}' if isinstance(val,float) else str(val),font=FONT_SM,fill='#243447',anchor='mm')
        d.text((x+bw/2,H-B+35),lab,font=FONT_SM,fill='#243447',anchor='mm')
    d.line((L,T,L,H-B),fill='#526577',width=2); d.line((L,H-B,W-R,H-B),fill='#526577',width=2)
    d.text((W/2,35),title,font=FONT_LG,fill=blue,anchor='mm'); d.text((12,T-40),ylabel,font=FONT_SM,fill=gray)
    im.save(path)
def line_chart(labels, values, title, ylabel, path, color):
    W,H=1500,700; L,R,T,B=140,50,90,120; im=Image.new('RGB',(W,H),'white'); d=ImageDraw.Draw(im)
    for k in range(6):
        y=T+(H-T-B)*k/5; val=100*(5-k)/5; d.line((L,y,W-R,y),fill='#d7e0e7',width=1); d.text((30,y-12),f'{val:.0f}%',font=FONT_SM,fill=gray)
    pts=[]; n=len(values)
    for i,(lab,val) in enumerate(zip(labels,values)):
        x=L+i*(W-L-R)/(n-1); y=H-B-(H-T-B)*val/100; pts.append((x,y)); d.text((x,H-B+35),lab,font=FONT_SM,fill='#243447',anchor='mm')
    d.line(pts,fill=color,width=6)
    for (x,y),val in zip(pts,values): d.ellipse((x-8,y-8,x+8,y+8),fill=color); d.text((x,y-30),f'{val:.1f}%',font=FONT_SM,fill='#243447',anchor='mm')
    d.line((L,T,L,H-B),fill='#526577',width=2); d.line((L,H-B,W-R,H-B),fill='#526577',width=2); d.text((W/2,35),title,font=FONT_LG,fill=blue,anchor='mm'); d.text((12,T-40),ylabel,font=FONT_SM,fill=gray); im.save(path)

names=[a['name'].replace('隨堂練習','練習') for a in assessment]
bar_chart(names,[a['n'] for a in assessment],'114-1 各次隨堂練習與作業繳交量','繳交紀錄數',ASSETS/'1141_submissions.png',[teal if '作業' not in n else orange for n in names])
bar_chart(names,[a['avg'] for a in assessment],'114-1 各次評量平均分數','平均分數',ASSETS/'1141_scores.png',[blue]*len(names),105)
line_chart([w['week'] for w in weekly],[w['rate'] for w in weekly],'114-2 十次學習單繳交率','名冊學生繳交率',ASSETS/'1142_completion.png',teal)
line_chart([w['week'] for w in selfweeks],[w['understand_rate'] for w in selfweeks],'114-2 各單元主觀理解程度','大致懂／很懂比例',ASSETS/'1142_understanding.png',blue)

def table(headers, body, widths=None):
    h='<table><thead><tr>'+''.join(f'<th>{esc(x)}</th>' for x in headers)+'</tr></thead><tbody>'
    for row in body: h+='<tr>'+''.join(f'<td>{x}</td>' for x in row)+'</tr>'
    return h+'</tbody></table>'
def img(name,caption): return f'<div class="figure"><img src="{(ASSETS/name).as_posix()}"><div class="caption">{esc(caption)}</div></div>'

assess_rows=[]
for a in assessment:
    assess_rows.append([esc(a['name']),str(a['n']),str(a['alt']),fmt(a['avg']),fmt(a['median']),fmt(pct(a['ge90'],a['n']))+'%',fmt(pct(a['ge60'],a['n']))+'%'])
weekly_rows=[[esc(w['week']),str(w['n']),fmt(w['rate'])+'%'] for w in weekly]
self_rows=[]
for w in selfweeks:
    c=w['ratings']; self_rows.append([esc(w['week']),str(w['n']),str(c.get('幾乎不懂',0)),str(c.get('有點懂',0)),str(c.get('大致懂',0)),str(c.get('很懂',0)),fmt(w['understand_rate'])+'%'])
likert_rows=[[f'Q{i+1}',fmt(x['mean'],2),fmt(x['positive'])+'%',esc(x['q'])] for i,x in enumerate(likert)]

avg_assess=mean([a['avg'] for a in assessment])
avg_weekly=mean([w['rate'] for w in weekly])
understand_total=pct(all_ratings.get('大致懂',0)+all_ratings.get('很懂',0),sum(all_ratings.values()))
atleast8=sum(x>=8 for x in completion_by_student)

css='''
@page { size:A4; margin:2cm 1.8cm 1.8cm 1.8cm; }
body { font-family:"Microsoft JhengHei","Noto Sans TC",sans-serif; color:#1f2937; font-size:10.5pt; line-height:1.55; }
h1 { color:#17365d; font-size:24pt; margin-top:0; } h2 { color:#1f4e79; font-size:17pt; border-bottom:2px solid #2a8c82; padding-bottom:5px; page-break-after:avoid; } h3 { color:#2f5f8f; font-size:13pt; page-break-after:avoid; }
p { margin:6px 0 9px; text-align:justify; } ul { margin-top:4px; } li { margin:4px 0; }
table { border-collapse:collapse; width:100%; margin:10px 0 16px; font-size:8.8pt; page-break-inside:auto; } th { background:#1f4e79; color:white; padding:6px; border:1px solid #cbd5e1; } td { padding:5px 6px; border:1px solid #cbd5e1; vertical-align:top; } tr:nth-child(even) td { background:#f5f8fb; }
.cover { text-align:center; padding-top:160px; page-break-after:always; } .cover .sub { color:#47637e; font-size:15pt; } .cover .date { margin-top:110px; color:#64748b; }
.callout { background:#eaf3f6; border-left:5px solid #2a8c82; padding:12px 15px; margin:12px 0; }
.kpis { width:100%; border-collapse:separate; border-spacing:8px; } .kpis td { background:#f0f6fa; border:1px solid #b9cedd; text-align:center; font-size:10pt; } .kpis strong { display:block; color:#17365d; font-size:18pt; }
.figure { text-align:center; margin:14px 0 18px; page-break-inside:avoid; } .figure img { width:92%; } .caption { color:#64748b; font-size:8.5pt; margin-top:3px; }
.pagebreak { page-break-before:always; } .small { font-size:8.5pt; color:#475569; } .toc li { margin:7px 0; }
'''

html_doc=f'''<!doctype html><html><head><meta charset="utf-8"><style>{css}</style></head><body>
<div class="cover"><div class="sub">教育部國民及學前教育署委辦課程</div><h1>Python 程式設計與 AI 實作<br>課程成果與學習成效結案報告</h1><div class="sub">114 學年度｜從生活入門到專案應用</div><p class="date">資料統整日期：中華民國 115 年 7 月 13 日<br>執行單位：國立臺灣師範大學科技應用與人力資源發展學系</p></div>

<h2>摘要</h2>
<p>本報告整合 <b>114-1 與 114-2</b> 兩學期之課程計畫、教材、隨堂練習、作業批改、修課名冊、單元學習單、期末作品、同儕互評與課程回饋，目的在呈現學生從 Python 基礎、資料處理與視覺化，逐步走向 AI 應用、專題製作與成果表達的學習歷程。分析單位以「有效繳交紀錄」及「有效問卷回覆」為主，個人姓名、電子郵件、學號與作品帳號均不在本報告揭露。</p>
<table class="kpis"><tr><td><strong>{total_1141:,}</strong>114-1 評量繳交紀錄</td><td><strong>{students}</strong>114-2 名冊學生</td><td><strong>{sum(w['n'] for w in weekly):,}</strong>114-2 學習單繳交</td><td><strong>{len(pdata)}</strong>期末互評回覆</td></tr><tr><td><strong>{n_feedback}</strong>114-1 課程回饋</td><td><strong>{project['with_both']}</strong>同時附程式與影片之留言</td><td><strong>{fmt(understand_total)}%</strong>單元自評大致懂／很懂</td><td><strong>{audit['audited_urls']}</strong>實際開啟稽核連結</td></tr></table>
<div class="callout"><b>主要判讀：</b>課程建立了高密度、可追蹤的實作證據，114-1 十三項評量共留下 {total_1141:,} 筆有效繳交紀錄；114-2 亦由 554 名學生名冊銜接至十次學習單。學生不只交程式，期末還需錄製影片、引用時間點與程式碼證據、評閱三件同儕作品並完成自我反思，顯示評量已從「能否執行」延伸到「能否解釋、判斷與改進」。</div>

<h2>一、分析範圍與方法</h2>
<h3>1.1 資料範圍</h3><p>本次盤點涵蓋資料夾內 144 個教學相關檔案（排除 .git 內部物件與作業系統暫存檔），包含 PDF、Word、Excel、Notebook、Markdown、Python、CSV、JSON、HTML、圖片與網址捷徑。核心量化來源為：<b>114-1成績總檔.xlsx</b>、<b>114-2 總表.xlsx</b>、114-1 課程回饋試算表，以及期末專題互評 Google 試算表。</p>
<h3>1.2 指標定義</h3><ul><li><b>繳交紀錄數：</b>每一工作表中具有可辨識分數或 V 註記的紀錄；同一學生重複繳交可能形成多筆，因此不直接等同獨立學生數。</li><li><b>格式異常：</b>提交 Notebook、文件、PDF 或無副檔名，而非指定 .py；內容仍由批改紀錄確認後給分。</li><li><b>正向回饋：</b>Likert 五點量表選擇 4 或 5。</li><li><b>理解較穩定：</b>單元自評選擇「大致懂」或「很懂」。</li><li><b>連結稽核：</b>開啟頁面並記錄 HTTP 狀態、重新導向與頁面標題；資料集欄位中的大量旅宿網址與 PTT 分頁採網域抽樣，不作為學生作品逐頁分析。</li></ul>
<h3>1.3 侷限</h3><p>資料夾包含 113-2 歷史回饋、114-1 完整成果與 114-2 進行中資料；本報告以可辨識的 114 學年度成果為主。114-2 計畫期程至 115 年 7 月 31 日，本次統計截至 115 年 7 月 13 日，故適合作為結案成果本文，但最終行政結報時仍應補入 7 月下旬新增之經費或活動憑證。評量檔沒有跨工作表穩定匿名識別碼，因此不能精確估計每位學生的縱向增分。</p>

<h2 class="pagebreak">二、課程設計與學習路徑</h2>
<p>課程定位為 2 學分、跨校遠距、實作與專題導向之多元選修。教材由 Python 基礎語法、資料結構、檔案與資料處理、資料視覺化，進階至線性迴歸、聊天機器人、Prompt／AI 助手、影像辨識、網路爬蟲、文字分類與自然語言處理。教學環境結合 YouTube 同步直播、Google Colab、GitHub 開放教材與線上批改機制。</p>
<p>學習證據形成四層遞進：</p><ol><li><b>即時理解：</b>隨堂練習與每週學習單。</li><li><b>技能應用：</b>五次個人實作作業與自動／人工複核批改。</li><li><b>整合創作：</b>以生活議題完成期末程式專題及影片。</li><li><b>反思與表達：</b>同儕互評、時間點證據、技術點摘錄、改進建議與自我反思。</li></ol>

<h2>三、114-1：作業提交與程式學習成果</h2>
<p>114-1 成績總檔共含 8 次隨堂練習與 5 次作業，合計 <b>{total_1141:,}</b> 筆具分數紀錄。各評量平均分數的簡單平均為 <b>{fmt(avg_assess)}</b> 分；其中 <b>{alt_1141}</b> 筆（{fmt(pct(alt_1141,total_1141))}%）屬非指定 .py 格式或 Notebook／文件格式。這些紀錄沒有被直接視為未完成，而是依內容正確性給予部分或完整分數，反映課程對初學者採取「先保留學習證據，再教導規格」的形成性評量策略。</p>
{img('1141_submissions.png','圖 1　114-1 各次評量之有效繳交紀錄數')}
{img('1141_scores.png','圖 2　114-1 各次評量之平均分數')}
{table(['評量','紀錄數','格式異常','平均','中位數','90分以上','60分以上'],assess_rows)}
<h3>3.1 學習歷程判讀</h3><ul><li>前四次隨堂練習留下 360–476 筆左右的高密度提交與複核紀錄，證明課堂活動確實轉化為可檢驗的程式輸出。</li><li>後段評量由基礎語法轉向資料處理、圖表、爬蟲與 AI 應用；作業量雖有波動，仍維持數百筆規模，形成跨週持續練習。</li><li>常見錯誤包括輸出格式不一致、程式逾時、語法錯誤、欄位處理與邏輯不符。這些錯誤已被批改檔具體記錄，適合回饋到下一輪教材與範例。</li><li>格式異常資料顯示「會寫但不會依規格提交」是初學階段的重要障礙；課程團隊以人工複核避免學生能力被檔案格式完全遮蔽。</li></ul>

<h2 class="pagebreak">四、114-2：參與持續度與單元理解</h2>
<p>114-2 名冊共 <b>{students}</b> 名學生、涵蓋 <b>{schools}</b> 所學校。十次學習單共記錄 <b>{sum(w['n'] for w in weekly):,}</b> 次繳交，平均每次 <b>{fmt(mean([w['n'] for w in weekly]))}</b> 份，按名冊計算平均繳交率為 <b>{fmt(avg_weekly)}%</b>。共有 <b>{atleast8}</b> 名學生至少完成 8 次（{fmt(pct(atleast8,students))}%）。</p>
{img('1142_completion.png','圖 3　114-2 十次學習單繳交率變化')}
{table(['日期','繳交人次','名冊繳交率'],weekly_rows)}
<p>前六次學習單由 460 份逐步降至 380 份，5 月份降到 178–224 份。這個下降幅度大於一般單週波動，應視為課程營運的重要訊號；可能因素包括段考、跨校行事曆、學期後段負荷、連結／帳號摩擦或教材難度增加，資料本身不足以單一歸因。</p>
<h3>4.1 主觀理解程度</h3><p>十個單元合計 <b>{sum(all_ratings.values()):,}</b> 份理解程度自評，其中「大致懂／很懂」占 <b>{fmt(understand_total)}%</b>，「幾乎不懂」占 <b>{fmt(pct(all_ratings.get('幾乎不懂',0),sum(all_ratings.values())))}%</b>。多數回覆落在「有點懂」與「大致懂」，顯示學生普遍能跟上核心概念，但尚未形成全面高自信的熟練度。</p>
{img('1142_understanding.png','圖 4　114-2 各單元「大致懂／很懂」比例')}
{table(['單元','回覆數','幾乎不懂','有點懂','大致懂','很懂','較穩定理解率'],self_rows)}

<h2 class="pagebreak">五、期末專題、作品表達與同儕互評</h2>
<p>期末成果留言彙整辨識到 <b>{project['comment_blocks']}</b> 個提交留言，其中 <b>{project['with_colab']}</b> 個附 Colab、<b>{project['with_video']}</b> 個附影片，<b>{project['with_both']}</b> 個同時提供程式與影片。作品內容從課堂技能延伸至生活議題，包括資料清理與視覺化、交通或旅宿資料分析、新聞爬蟲與摘要、情緒／文字分析、線性迴歸預測、圖像處理、聊天機器人與生活管理工具。</p>
<p>互評表取得 <b>{len(pdata)}</b> 份有效回覆，學生被要求觀摩三件作品，提供兩個影片時間點證據、至少一項程式細節、優點、下一步建議與 1–5 分評分；實際蒐集 <b>{len(peer_scores)}</b> 個有效互評分數，平均 <b>{fmt(mean(peer_scores),2)}</b> 分，中位數 <b>{fmt(med(peer_scores),1)}</b> 分。另有 <b>{reflection_complete}</b> 份回覆完整填寫技術學習、重做改進、困難解法與給下一屆建議等反思欄位。</p>
<div class="callout"><b>成效意義：</b>「程式碼 + 影片 + 證據式互評」使成果不只是提交檔案，而是要求學生把操作知識轉成可說明的概念，並練習看懂他人的程式、提出具體回饋與形成下一步改進。這同時對應技術實作、資訊判讀、口語表達與後設認知。</div>

<h2>六、學生回饋與課程感受</h2>
<p>114-1 共回收 <b>{n_feedback}</b> 份回饋。十項量化題之平均介於 <b>{fmt(min(x['mean'] for x in likert),2)}</b> 至 <b>{fmt(max(x['mean'] for x in likert),2)}</b> 分，整體十題平均 <b>{fmt(mean([x['mean'] for x in likert]),2)}</b> 分；各題正向回饋率介於 <b>{fmt(min(x['positive'] for x in likert))}%</b> 至 <b>{fmt(max(x['positive'] for x in likert))}%</b>。學生對 AI 與生活問題連結、實際模型應用、教學與實作比例及進一步學習意願均留下可量化證據。</p>
{table(['題號','平均（5分）','正向率','題意'],likert_rows)}
<p>互動面仍有改善空間：質性問卷中，<b>{interaction.get('沒留言過',0)}</b> 人表示未曾與老師留言互動；在 Play／Share 階段，<b>{post_freq.get('從沒發言過',0)}</b> 人表示從未發文。這表示大量作業繳交不必然轉化為公開討論參與，後續應以更低壓、結構化的互動任務提升共學感。</p>

<h2 class="pagebreak">七、連結與數位成果稽核</h2>
<p>所有可解析文件共辨識 <b>{audit['unique_urls']:,}</b> 個正規化唯一網址。實際開啟 <b>{audit['audited_urls']}</b> 個教學、作品、Colab、YouTube、Google 文件、GitHub、表單與工具連結；其中 HTTP 200 為 <b>{audit['status_counts'].get('200',0)}</b> 個。另有 <b>{audit['bulk_dataset_urls_not_individually_fetched']:,}</b> 個屬旅宿資料欄位或 PTT 分頁型資料值，為避免對資料來源網站造成大量請求，採網域與代表頁檢查，不視為作品連結逐頁分析。</p>
<p>連結稽核發現部分 GitHub 教材舊路徑回傳 404，若學生仍從舊 PDF／Notebook 點擊，會增加學習摩擦；Google 表單或文件亦有需登入（401）或權限不一致情形。另一方面，多數 Colab 與 YouTube 連結可到達平台頁面，但「平台頁可開啟」不等於內容對所有人具有檢視權限，建議結案封存前以無痕視窗再做一次權限抽查。</p>

<h2>八、整體成效判斷</h2>
<ol><li><b>規模可見：</b>兩學期累積數千筆作業／學習單紀錄，顯示跨校遠距課程具有實際參與，而非僅提供教材。</li><li><b>歷程可見：</b>批改檔保留分數、執行狀態、錯誤類型與時間；114-2 再加入逐單元理解自評，使「做了什麼、卡在哪裡」能被追蹤。</li><li><b>能力可見：</b>評量從語法、資料結構與檔案處理，推進到資料分析、視覺化、AI／NLP 應用與專題整合。</li><li><b>表達可見：</b>期末影片與證據式互評要求學生解釋程式、引用時間點、辨識技術細節並提出改進。</li><li><b>公平支持可見：</b>格式錯誤仍經人工複核，降低單一提交規格對初學者成績的誤傷；全國跨校線上模式也擴大 AI 實作資源可近性。</li></ol>
<p>綜合而言，本課程效果最強的證據不是單一平均分數，而是形成了「大量實作—錯誤回饋—反覆提交—專題整合—同儕互評—自我反思」的完整學習閉環。資料亦誠實呈現兩個待改善面向：學期後段繳交下降，以及公開互動／留言參與偏低。</p>

<h2>九、後續改善建議</h2>
<ol><li><b>建立穩定匿名學習者 ID：</b>跨工作表使用相同代碼，才能做真正的個人縱向學習曲線、留存率與增值分析。</li><li><b>設置學期後段預警：</b>連續兩週未交、繳交率快速下降或「幾乎不懂」者，由合作教師／助教進行低負擔提醒與補交支持。</li><li><b>把檔案規格納入第一週微任務：</b>用 10 分鐘完成檔名、副檔名、分享權限與 Colab 匯出練習，減少「會做但交錯格式」。</li><li><b>降低互動門檻：</b>將每週留言改為固定句型（我完成了／我卡在／我可以幫忙），並讓學生先在校內小組回覆再進全班公開區。</li><li><b>維護連結清單：</b>每學期開始前自動檢查 GitHub 404、Google 權限與表單狀態；教材以單一穩定入口連到當期版本。</li><li><b>補強成效設計：</b>增加前後測或共同錨題，以區分「高分來自題目容易」與「能力確實成長」，並保留同一能力面向的前後可比題。</li></ol>

<h2 class="pagebreak">附錄 A　主要資料來源</h2>
<ul><li>114-1成績總檔.xlsx：8 次隨堂練習、5 次作業、分數、執行狀態、錯誤與格式複核。</li><li>114-2 總表.xlsx：554 名學生名冊、十次繳交總表、十個單元學習單與理解自評。</li><li>114-1 課程回饋試算表：95 份量化與質性回饋。</li><li>114-1 期末小專題影片觀摩與互評學習單：作品、三件同儕互評、時間點／程式碼證據與反思。</li><li>New Text Document.txt：期末專題留言與作品連結彙整。</li><li>PythonAI4Beginners：10 個單元 Notebook、18 週示例教材、作業說明、參考解答、練習與資料集。</li><li>課程計畫書、課程大綱、行政函文、經費與歷史成果文件：課程目標、執行方式、評量設計與計畫脈絡。</li></ul>
<h2>附錄 B　資料治理與解讀提醒</h2><ul><li>本報告不列學生姓名、電子郵件、學號、帳號或個別作品網址。</li><li>「繳交紀錄」不等同「獨立學生」，不可直接當作修課人數。</li><li>自評反映主觀理解，應與作業表現、前後測及教師觀察共同解讀。</li><li>114-2 行政計畫期程尚未到 115 年 7 月 31 日，最終送件前應確認是否有新增成果與經費資料。</li><li>網路連結狀態為 115 年 7 月 13 日稽核快照，後續可能因權限、刪除或平台政策改變。</li></ul>
</body></html>'''

(WORK/'.codex_high_school_114_report.html').write_text(html_doc,encoding='utf-8')
stats={'grade_file':str(grade_path),'roster_file':str(roster_path),'feedback_file':str(feedback_path),'assessment':assessment,'total_1141':total_1141,'alt_1141':alt_1141,'students_1142':students,'schools_1142':schools,'weekly':weekly,'selfweeks':selfweeks,'all_ratings':dict(all_ratings),'feedback_n':n_feedback,'likert':likert,'project':project,'peer_n':len(pdata),'peer_scores_n':len(peer_scores),'peer_score_mean':mean(peer_scores),'reflection_complete':reflection_complete,'audit':audit}
(WORK/'.codex_high_school_114_stats.json').write_text(json.dumps(stats,ensure_ascii=False,indent=2),encoding='utf-8')
print(json.dumps({'html':str(WORK/'.codex_high_school_114_report.html'),'stats':str(WORK/'.codex_high_school_114_stats.json'),'peer_n':len(pdata),'total_1141':total_1141,'students_1142':students},ensure_ascii=False))
