import csv, json, os, re, sys, zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import fitz
import openpyxl

ROOT = Path(r"C:\GitHub\high-school\114")
OUT = Path(r"C:\GitHub\fih-seattle.github.io\.codex_high_school_114_extract.json")
URL_RE = re.compile(r"https?://[^\s<>\]\[\)\(\}\{\"']+", re.I)

def clean(s):
    return re.sub(r"\n{3,}", "\n\n", str(s).replace("\x00", "")).strip()

def xml_text(path):
    with zipfile.ZipFile(path) as z:
        chunks=[]
        for name in z.namelist():
            if name.startswith("word/") and name.endswith(".xml"):
                try:
                    root=ET.fromstring(z.read(name))
                    chunks.extend((el.text or "") for el in root.iter() if el.tag.endswith('}t'))
                except Exception: pass
        return "\n".join(chunks)

def xlsx_text(path):
    wb=openpyxl.load_workbook(path, data_only=False, read_only=False)
    out=[]
    for ws in wb.worksheets:
        out.append(f"\n### SHEET: {ws.title} ({ws.max_row}x{ws.max_column})")
        for row in ws.iter_rows():
            vals=[]
            for c in row:
                v=c.value
                if v is not None: vals.append(f"{c.coordinate}={v}")
                if c.hyperlink and c.hyperlink.target: vals.append(f"{c.coordinate}_LINK={c.hyperlink.target}")
            if vals: out.append(" | ".join(vals))
    return "\n".join(out)

def ipynb_text(path):
    d=json.loads(path.read_text(encoding='utf-8'))
    out=[]
    for i,c in enumerate(d.get('cells',[]),1):
        src=''.join(c.get('source',[]))
        out.append(f"\n### CELL {i} [{c.get('cell_type','')}]\n{src}")
        for o in c.get('outputs',[]):
            txt=o.get('text') or o.get('data',{}).get('text/plain')
            if txt: out.append(''.join(txt) if isinstance(txt,list) else str(txt))
    return '\n'.join(out)

def read_file(p):
    ext=p.suffix.lower()
    if ext=='.pdf':
        doc=fitz.open(p); return '\n'.join(page.get_text() for page in doc)
    if ext=='.docx': return xml_text(p)
    if ext=='.xlsx': return xlsx_text(p)
    if ext=='.ipynb': return ipynb_text(p)
    if ext in {'.md','.txt','.py','.css','.html','.json','.gitignore','.url'}:
        for enc in ('utf-8','utf-8-sig','cp950','big5','latin1'):
            try: return p.read_text(encoding=enc)
            except Exception: pass
    if ext=='.csv':
        for enc in ('utf-8-sig','cp950','big5','utf-8'):
            try: return p.read_text(encoding=enc)
            except Exception: pass
    if ext=='.ods':
        with zipfile.ZipFile(p) as z:
            root=ET.fromstring(z.read('content.xml'))
            return '\n'.join((e.text or '') for e in root.iter() if e.text)
    return ''

records=[]
for p in sorted(ROOT.rglob('*')):
    if not p.is_file() or '.git' in p.parts or p.name=='.DS_Store': continue
    rec={'path':str(p.relative_to(ROOT)), 'ext':p.suffix.lower(), 'size':p.stat().st_size}
    try:
        text=clean(read_file(p))
        rec['text']=text
        rec['urls']=sorted(set(URL_RE.findall(text)))
        rec['status']='ok' if text else 'binary_or_empty'
    except Exception as e:
        rec['text']=''; rec['urls']=[]; rec['status']=f'error: {type(e).__name__}: {e}'
    records.append(rec)

OUT.write_text(json.dumps({'root':str(ROOT),'records':records},ensure_ascii=False,indent=2),encoding='utf-8')
print(f"records={len(records)} output={OUT}")
print("statuses", {s:sum(r['status']==s for r in records) for s in sorted(set(r['status'] for r in records))})
print("urls", sum(len(r['urls']) for r in records), "unique", len(set(u for r in records for u in r['urls'])))
