import json, re, shutil, sys, unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import fitz
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')
BOOK = Path(r'C:\GitHub\scores\第一階段書審評分表-蔡芸琤.xlsx')
PDFDIR = Path(r'C:\GitHub\scores\materials\PDF')
VIDDIR = Path(r'C:\GitHub\scores\materials\Videos')
REPORT = Path(r'C:\GitHub\scores\materials\review_candidates.json')

def norm(s):
    s = unicodedata.normalize('NFKC', str(s or '')).lower()
    return re.sub(r'[^0-9a-z\u4e00-\u9fff]+', '', s)

def all_text(path):
    d = fitz.open(path)
    return '\n'.join(p.get_text() for p in d)

def has_any(text, terms):
    return any(t.lower() in text.lower() for t in terms)

wb = load_workbook(BOOK, data_only=False)
entries = []
for sn in ('大專組', '高中職組'):
    ws = wb[sn]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 3).value:
            entries.append({'sheet':sn,'row':row,'team':str(ws.cell(row,1).value),
                            'title':str(ws.cell(row,2).value),'case':str(ws.cell(row,3).value)})

pdfs = []
for p in PDFDIR.glob('*.pdf'):
    text = all_text(p)
    pdfs.append({'path':p,'name':p.name,'text':text,'pages':fitz.open(p).page_count})

# Greedy global matching using both filename and document opening text.
pairs = []
for ei,e in enumerate(entries):
    target_team, target_title = norm(e['team']), norm(e['title'])
    for pi,p in enumerate(pdfs):
        hay = norm(p['name'] + p['text'][:6000])
        exact = (35 if target_team and target_team in hay else 0) + (45 if target_title and target_title in hay else 0)
        sim = 20 * max(SequenceMatcher(None,target_team,hay[:max(1,len(target_team)*4)]).ratio(),
                       SequenceMatcher(None,target_title,hay[:max(1,len(target_title)*4)]).ratio())
        pairs.append((exact+sim,ei,pi))
assigned_e, assigned_p, match = set(), set(), {}
for score,ei,pi in sorted(pairs,reverse=True):
    if ei not in assigned_e and pi not in assigned_p:
        assigned_e.add(ei); assigned_p.add(pi); match[ei]=(pi,score)

records=[]
for ei,e in enumerate(entries):
    pi,match_score=match[ei]; p=pdfs[pi]; t=p['text']; tl=t.lower()
    # Completeness: the official template's six substantive fields.
    fields = {
      'motivation': has_any(t,['計畫動機','計劃動機','提案動機']),
      'objectives': has_any(t,['計畫目標','計劃目標','提案目標']),
      'execution': has_any(t,['執行內容','實施內容','系統架構','計畫內容']),
      'schedule': has_any(t,['時程','甘特圖','工作進度','執行期程']),
      'benefit': has_any(t,['預期效益','預期成果','社會影響','效益評估']),
      'sdg': has_any(t,['sdg','永續發展目標']),
    }
    completeness=max(0,30-5*sum(not v for v in fields.values()))

    # Technical implementation (12) + Demo Day presentation readiness (8).
    cats = [
      has_any(t,['python','tensorflow','pytorch','yolo','機器學習','深度學習','影像辨識','自然語言']),
      has_any(t,['app','網站','平台','前端','後端','資料庫','api','雲端']),
      has_any(t,['esp32','arduino','感測器','iot','物聯網','鏡頭','硬體']),
      has_any(t,['政府開放資料','開放資料','資料集','dataset','中央氣象署','環境部']),
    ]
    evidence = sum([has_any(t,['原型','prototype','實作成果','已完成','測試結果']),
                    has_any(t,['準確率','辨識率','實驗','驗證','測試']),
                    has_any(t,['系統架構','流程圖','介面','截圖'])])
    impl=min(12,5+sum(cats)+evidence)
    case=e['case']
    if case=='1151121-91': demo=0; video_status='unplayable'
    elif (VIDDIR/f'{case}.pdf').exists(): demo=2; video_status='slides_only'
    else: demo=8; video_status='playable'
    feasibility=impl+demo

    # Innovation: originality, problem solving, and sustainability, grounded in full text.
    local = has_any(t,['南投','埔里','竹山','草屯','仁愛','信義','山城'])
    problem = has_any(t,['痛點','問題','需求','挑戰','風險'])
    integration = sum(cats)>=2
    novel = has_any(t,['創新','首創','不同於','差異化','獨特'])
    sustainable = has_any(t,['永續','循環','節能','減碳','韌性','長期營運','商業模式'])
    # Reserve the top three points for genuinely demonstrated originality,
    # rather than awarding full marks merely for mentioning innovation.
    innovation=min(25,10+2*local+2*problem+3*integration+2*novel+3*sustainable)

    # Social impact: SDGs must be stated and connected to concrete plan content.
    sdgs=sorted(set(re.findall(r'(?:sdgs?|SDGs?)\s*[-：:]?\s*(\d{1,2})',t,re.I)))
    impact=25
    if not fields['sdg']: impact=15
    elif not local: impact-=3
    if not fields['benefit']: impact-=4
    if len(sdgs)>6: impact-=2  # unusually broad claims are less well substantiated
    impact=max(0,impact)
    records.append({**e,'pdf':p['name'],'pages':p['pages'],'match_score':round(match_score,1),
                    'fields':fields,'sdgs':sdgs,'video_status':video_status,
                    'scores':{'F':feasibility,'G':innovation,'H':completeness,'I':impact,
                              'J':feasibility+innovation+completeness+impact},
                    'signals':{'tech_categories':sum(cats),'evidence':evidence,'chars':len(t)}})

REPORT.write_text(json.dumps(records,ensure_ascii=False,indent=2),encoding='utf-8')
print(json.dumps([{k:r[k] for k in ('sheet','row','case','team','title','pdf','pages','match_score','fields','sdgs','video_status','scores','signals')} for r in records],ensure_ascii=False,indent=2))

if '--write' in sys.argv:
    backup=BOOK.with_name(BOOK.stem+'-填分前備份-'+datetime.now().strftime('%Y%m%d-%H%M%S')+BOOK.suffix)
    shutil.copy2(BOOK,backup)
    for r in records:
        ws=wb[r['sheet']]
        for col in 'FGHIJ': ws[f'{col}{r["row"]}']=r['scores'][col]
    wb.save(BOOK)
    print('BACKUP',backup)
    print('SAVED',BOOK)
